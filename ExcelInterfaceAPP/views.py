from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.core.paginator import Paginator

from .forms import *
from .models import *
from django.core.exceptions import ValidationError
from django.core.cache import cache
import pandas as pd
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Excel, CheckboxStatus
from .forms import ExcelUploadForm


def home(request):
    q = request.GET.get('q')
    aba_selecionada = request.GET.get('aba')
    form = ExcelUploadForm()
    erro = None
    excel_dados = None
    abas_disponiveis = []
    abas_filtradas = []
    abas_sem_resultado = []
    saved_checkboxes = set()
    departamento = request.GET.get("departamento")
    tipo_aba = None

    DEPARTAMENTO_COLUNAS_DADOS = {
        "Projetos": ["CHEGADA DO AÇO"],
        "Pcp": ["CHEGADA DO AÇO"],
        "Compras": ["CHEGADA DO AÇO"],
        "CAM": ["CHEGADA DO AÇO"],
        "Processos": ["CHEGADA DO AÇO", "PROGRAMA NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
        "Produção": ["CHEGADA DO AÇO", "PROGRAMA NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
    }

    DEPARTAMENTO_COLUNAS_M = {
        "Projetos": ["aço"],
        "Pcp": ["Modelo 3d", "aço"],
        "Compras": ["Modelo 3d", "aço"],
        "CAM": ["Modelo 3d", "aço"],
        "Processos": ["Modelo 3d", "aço", "Programa NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
        "Produção": ["Modelo 3d", "aço", "Programa NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
    }

    COLUNAS_FIXAS = {
        "DADOS": ["Name.1", "Name.2.1", "STATUS"],
        "M": ["Ord Serv", "Item", "Descrição"],
    }

    def limpar_checkboxes_em_colunas(df, colunas):
        df = df.copy()
        for coluna in colunas:
            if coluna in df.columns:
                df[coluna] = df[coluna].apply(lambda x: "" if str(x).strip() == "☑" else x)
        return df

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "excel":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                arquivo_excel = form.cleaned_data["arquivo"]
                try:
                    extensoes_validas = ['.xls', '.xlsx', 'xlsm']
                    if not any(arquivo_excel.name.endswith(ext) for ext in extensoes_validas):
                        raise ValidationError("O arquivo deve ser um Excel (.xls ou .xlsx).")

                    excel_obj = Excel(nome=arquivo_excel.name, arquivo=arquivo_excel)
                    excel_obj.save()
                    cache.delete('excel_dfs')  # Limpa cache para recarregar depois
                except ValidationError as e:
                    erro = str(e)
                except Exception as e:
                    erro = f"Erro ao processar o arquivo: {e}"

        elif acao == "salvar_checkboxes":
            aba_checkbox = request.POST.get("aba")
            checked_data = request.POST.getlist("checkbox")

            for item in checked_data:
                try:
                    linha_idx, coluna = item.split("__")
                    CheckboxStatus.objects.update_or_create(
                        aba=aba_checkbox,
                        linha_index=int(linha_idx),
                        coluna=coluna,
                        defaults={"checked": True}
                    )
                except Exception as e:
                    erro = f"Erro ao salvar checkbox: {e}"

    try:
        # 1. Tentar pegar do cache
        dfs_por_aba = cache.get('excel_dfs', {})
        ultimo_excel_data = Excel.objects.order_by("-data").values_list("data", flat=True).first()
        cache_data_key = cache.get("excel_cache_data")

        # 2. Se mudou o arquivo mais recente, limpar cache para recarregar
        if ultimo_excel_data and cache_data_key != str(ultimo_excel_data):
            cache.delete("excel_dfs")
            cache.set("excel_cache_data", str(ultimo_excel_data))
            dfs_por_aba = {}

        # 3. Se cache vazio, carregar todos os arquivos do banco
        if not dfs_por_aba:
            todos_excel = Excel.objects.all().order_by('data')
            dfs_por_aba = {}

            for excel in todos_excel:
                try:
                    caminho_arquivo = excel.arquivo.path
                    excel_dict = pd.read_excel(caminho_arquivo, sheet_name=None)

                    for aba_nome, df in excel_dict.items():
                        df_limpo = df.fillna("").loc[:, ~df.columns.str.contains('^Unnamed')]
                        nome_completo = aba_nome

                        if nome_completo in dfs_por_aba:
                            try:
                                if df_limpo.equals(dfs_por_aba[nome_completo]):
                                    continue  # já existe igual
                                else:
                                    sufixo = 1
                                    novo_nome = f"{nome_completo} ({sufixo})"
                                    while novo_nome in dfs_por_aba:
                                        sufixo += 1
                                        novo_nome = f"{nome_completo} ({sufixo})"
                                    nome_completo = novo_nome
                            except:
                                continue

                        dfs_por_aba[nome_completo] = df_limpo

                except Exception as e:
                    erro = f"Erro ao processar o arquivo '{excel.nome}': {e}"

            cache.set('excel_dfs', dfs_por_aba, timeout=300)

        # 4. AGORA APLICAR A LÓGICA DE CONEXÃO ENTRE ABAS M e DADOS
        for aba_m, df_m in list(dfs_por_aba.items()):
            if aba_m.startswith("M") and aba_m != "MOLDES EM PROGRESSO":
                sufixo = aba_m[1:]  # tudo após o "M"
                aba_dados = f"DADOS {sufixo}"

                df_dados = dfs_por_aba.get(aba_dados)
                if df_dados is not None:

                    coluna_chave = "Item"  # ajuste para sua chave real
                    col_status = "STATUS"  # ajuste para sua coluna de status

                    if coluna_chave in df_m.columns and coluna_chave in df_dados.columns and col_status in df_dados.columns:

                        df_merge = df_m.merge(df_dados[[coluna_chave, col_status]], on=coluna_chave, how="left")
                        df_merge["Feito"] = df_merge[col_status].apply(lambda x: "☑" if pd.notna(x) and str(x).strip() else "")

                        dfs_por_aba[aba_m] = df_merge

        # 5. Segue seu fluxo de filtragem, paginação, etc...
        abas_disponiveis = list(dfs_por_aba.keys())
        dfs_filtrados = []

        if q:
            dfs_filtrados = {}
            for nome_aba, df_original in dfs_por_aba.items():
                mask = df_original.astype(str).apply(lambda col: col.str.contains(q, case=False, na=False)).any(axis=1)
                df_filtrado = df_original[mask]
                dfs_filtrados[nome_aba] = df_filtrado

                if df_filtrado.empty:
                    abas_sem_resultado.append(nome_aba)
                else:
                    abas_filtradas.append(nome_aba)

            aba = aba_selecionada or (abas_filtradas[0] if abas_filtradas else None)
            if aba:
                df = dfs_filtrados.get(aba, pd.DataFrame())
            else:
                erro = "Nenhum resultado encontrado."
                df = pd.DataFrame()
        else:
            # Corrigido: escolha uma aba com dados
            aba = aba_selecionada
            if not aba:
                for nome_aba, df_test in dfs_por_aba.items():
                    if not df_test.empty:
                        aba = nome_aba
                        break
                if not aba:
                    erro = "Nenhuma aba com dados disponíveis."
                    df = pd.DataFrame()
            df = dfs_por_aba.get(aba, pd.DataFrame()) if aba else pd.DataFrame()

        if df is None or df.empty:
            erro = "Nenhum resultado encontrado."
            df = pd.DataFrame()

        if aba and aba.startswith("M"):
            colunas_checkbox = ["Modelo 3d", "aço", "Programa NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"]
            df = limpar_checkboxes_em_colunas(df, colunas_checkbox)

        df = df.replace("☑", "")

        tipo_aba = None
        if aba:
            if aba.startswith("DADOS"):
                tipo_aba = "DADOS"
            elif aba.startswith("M"):
                tipo_aba = "M"

        if not df.empty and tipo_aba and departamento:
            colunas_fixas = COLUNAS_FIXAS.get(tipo_aba, [])
            colunas_departamento = (
                DEPARTAMENTO_COLUNAS_DADOS.get(departamento, []) if tipo_aba == "DADOS"
                else DEPARTAMENTO_COLUNAS_M.get(departamento, [])
            )

            colunas_desejadas = [col for col in df.columns if col in colunas_fixas or col in colunas_departamento]
            df = df[colunas_desejadas]

        if not df.empty:
            lista_de_registros = df.to_dict(orient="records")
            paginator = Paginator(lista_de_registros, 50)
            page_number = request.GET.get("page")
            page_obj = paginator.get_page(page_number)
            excel_dados = page_obj

        checkboxes = CheckboxStatus.objects.filter(aba=aba).values_list("linha_index", "coluna", flat=False) if aba else []
        saved_checkboxes = set(f"{linha}__{col}" for linha, col in checkboxes)

    except Excel.DoesNotExist:
        erro = "Nenhum arquivo Excel foi enviado ainda."
    except Exception as e:
        erro = f"Erro ao carregar o arquivo: {e}"

    if 'tipo_aba' not in locals():
        tipo_aba = None

    contexto = {
        "erro": erro,
        "ExcelDados": excel_dados,
        "abas": abas_disponiveis,
        "abas_sem_resultado": abas_sem_resultado,
        "aba_selecionada": aba,
        "form": form,
        "q": q,
        "saved_checkboxes": saved_checkboxes,
        "colunas_checkbox_m": ["Modelo 3d", "aço", "Programa NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
        "colunas_checkbox_d": ["STATUS", "Programa NC", "DMG", "600 II", "800 II", "600 I", "800 I", "IXION II"],
        "departamento": departamento,
        "colunas_visiveis": df.columns if not df.empty else [],
        "tipo_aba": tipo_aba,
        "departamentos": list(DEPARTAMENTO_COLUNAS_DADOS.keys()),
    }

    return render(request, 'ExcelInterface/table.html', contexto)

def ChecklistPagina(request):
    checklists = Checklist.objects.all()
    return render(request, 'ExcelInterface/checklist.html', {'checklists': checklists})


def get_abas_do_excel(file_path):
    try:
        excel = pd.ExcelFile(file_path)
        return excel.sheet_names
    except Exception as e:
        return []


def criar_checklist(request):
    abas = []
    excel_objs = Excel.objects.all().order_by('data')

    try:
        for excel_obj in excel_objs:
            caminho_arquivo = excel_obj.arquivo.path
            abas_do_arquivo = get_abas_do_excel(caminho_arquivo)
            for aba in abas_do_arquivo:
                if aba not in abas:
                    abas.append(aba)
    except Exception as e:
        messages.error(request, f"Erro ao carregar abas do Excel: {e}")

    if request.method == 'POST':
        form = CheckListForm(request.POST, abas=abas)
        if form.is_valid():
            checklist = form.save(commit=False)
            if excel_objs.exists():
                checklist.excel = excel_objs.last()  # ou use outro critério para escolher o Excel vinculado
            checklist.save()

            itens_json = request.POST.get('itens_json', '[]')
            try:
                itens = json.loads(itens_json)
                for item in itens:
                    ChecklistITEM.objects.create(
                        Lista=checklist,
                        nome_item=item.get('nome', ''),
                        descricaoItem=item.get('descricao', ''),
                    )
            except json.JSONDecodeError:
                messages.error(request, "Erro ao processar os itens adicionados.")

            messages.success(request, f"Checklist '{checklist.nome_lista}' criada com sucesso!")
            return redirect('Checklist')
    else:
        form = CheckListForm(abas=abas)

    return render(request, 'ExcelInterface/Criarchecklist.html', {'form': form})
    
    
def atualizar_status_item(request):
    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        novo_status = request.POST.get('status') == 'true'

        try:
            item = ChecklistITEM.objects.get(id=item_id)
            item.statusItem = novo_status
            item.save()
            return JsonResponse({'success': True})
        except ChecklistITEM.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Item não encontrado'})

    return JsonResponse({'success': False, 'error': 'Método inválido'})


def DeletarLista(request, item_type, objetoId):
    if request.method == "POST":
        item = get_object_or_404(Checklist, pk=objetoId)
        item.delete()
        return JsonResponse({'message': 'Objeto excluído com sucesso.'}, status=200)
    return JsonResponse({'error': 'Método não permitido'}, status=405)


def abrir_checklist(request, checklist_id):
    checklist = get_object_or_404(Checklist, id=checklist_id)
    html = render(request, 'partials/checklist_detail.html', {'checklist': checklist}).content.decode('utf-8')
    return JsonResponse({'html': html})



def exportar_todas_planilhas(request):
    cache.delete('excel_dfs')
    excel_objs = Excel.objects.all().order_by('data')
    checkboxes_marcados = request.session.get("checkboxes_marcados", [])

    wb = Workbook()
    wb.remove(wb.active)

    for excel_obj in excel_objs:
        caminho_arquivo = excel_obj.arquivo.path
        try:
            excel_dict_raw = pd.read_excel(caminho_arquivo, sheet_name=None, header=0)
        except Exception:
            continue  # Ignora arquivos com erro de leitura

        excel_dict = {}

        for nome_aba, df in excel_dict_raw.items():
            if df.empty:
                continue

            # Manter somente colunas cujo nome é válido (não vazio e não unnamed)
            colunas_validas = [col for col in df.columns if str(col).strip() != "" and not str(col).lower().startswith("unnamed")]
            df = df[colunas_validas]

            excel_dict[nome_aba] = df

        for nome_aba, df in excel_dict.items():
            ws = wb.create_sheet(title=nome_aba[:31])
            colunas = list(df.columns)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="4F81BD")
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="center")

                        # Marcação de checkbox
                        linha_idx = r_idx - 2
                        coluna_nome = colunas[c_idx - 1]
                        chave = f"{linha_idx}__{coluna_nome}"
                        if chave in checkboxes_marcados:
                            cell.value = f"☑ {cell.value}"

                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'),
                    )

            for column_cells in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 5
                ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="TodasPlanilhasFormatadas.xlsx"'
    wb.save(response)
    return response



def importar_planilha(request):
    if request.method == 'POST' and request.FILES.get('planilha'):
        arquivo = request.FILES['planilha']
        excel = Excel.objects.create(arquivo=arquivo)

        return redirect('/')

    return redirect('/') 